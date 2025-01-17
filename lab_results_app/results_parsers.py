"""
Laboratory Results Parser and Database Integration Module

This module provides functionality for parsing and processing laboratory test results
from different experimental methods (TNS and Zeta Potential) and storing them in a SQLite database.
It handles file parsing, data validation, statistical calculations, and database operations.

Supported Result Types:
- TNS (Transfection Normalized to Standard): Excel-based results with specific cell formatting
- Zeta Potential: CSV-based results with measurements of particle surface charge

Dependencies:
    - openpyxl: For Excel file handling
    - polars: For efficient data manipulation
    - sqlite3: For database operations
    - statistics: For statistical calculations
    - datetime: For timestamp handling
"""

from openpyxl import load_workbook
import polars as pl
import polars.datatypes
import polars.selectors as cs
import statistics
from datetime import datetime
from typing import Tuple, List

from . import database_helpers

class ResultsParsingError(Exception): pass

def parse_file(filename: str) -> List[Tuple[str, float]]:
    """
    Determines the type of result file and routes it to the appropriate parser.

    The function attempts to read the file as a text file first. If it succeeds,
    it's considered a Zeta Potential file (CSV). If it fails with UnicodeDecodeError,
    it's considered a TNS file (Excel).

    Args:
        filename: Path to the result file to be parsed

    Returns:
        List of tuples containing (formulation_id, calculated_value)

    Raises:
        ResultsParsingError: If the file type cannot be determined
        UnicodeDecodeError: If the file cannot be read as text
    """
    type = None
    try:
        with open(filename, newline='') as f:
            f.read(1024)
        type = "ZETA_POTENTIAL"
    except UnicodeDecodeError:
        type = "TNS"
    
    if type == "TNS":
        return parse_tns_results(filename)
    if type == "ZETA_POTENTIAL":
        return parse_zeta_potential_results(filename)
    raise ResultsParsingError("Could not determine filetype!")


# Constants defining the expected structure of TNS result files
TNS_HEADER = ('<>', 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)

TNS_SCHEMA = {
    "<>": polars.datatypes.String,
    "1": polars.datatypes.UInt64,
    "2": polars.datatypes.UInt64,
    "3": polars.datatypes.UInt64,
    "4": polars.datatypes.UInt64,
    "5": polars.datatypes.UInt64,
    "6": polars.datatypes.UInt64,
    "7": polars.datatypes.UInt64,
    "8": polars.datatypes.UInt64,
    "9": polars.datatypes.UInt64,
    "10": polars.datatypes.UInt64,
    "11": polars.datatypes.UInt64,
    "12": polars.datatypes.UInt64,
}

def parse_tns_results(filename: str) -> List[Tuple[str, float]]:
    """
    Parses TNS (Transfection Normalized to Standard) results from an Excel file.

    The TNS file must follow a specific format:
    - Dimensions must be A1:M11
    - Cell A1 must contain "Instrument:"
    - Cell B1 must contain the instrument name
    - Data must start with '<>' marker
    - Contains 3 sets of triplicate measurements
    - Contains control values in columns 10-12

    Args:
        filename: Path to the Excel file containing TNS results

    Returns:
        List of tuples containing (formulation_id, calculated_value)

    Raises:
        ResultsParsingError: If the file format is invalid or if calculated values are below threshold
    """
    wb = load_workbook(filename=filename)
    ws = wb.active

    # Validate worksheet dimensions
    if ws.dimensions != "A1:M11":
        raise ResultsParsingError(
            "Invalid data format. The extent of occupied cells should be A1:M11"
        )

    data = list(ws.values)

    # Validate instrument information
    if data[0][0] != "Instrument:":
        raise ResultsParsingError("Invalid data format. 'Instrument:' label is missing")

    instrument = data[0][1]
    if not instrument:
        raise ResultsParsingError("Invalid data format. Missing instrument name")

    # Validate empty cells
    if any(data[0][2:] + data[1]):
        raise ResultsParsingError(
            "Invalid data format. Cells that should be empty aren't")

    # Validate data header
    if data[2][0] != "<>":
        raise ResultsParsingError(
            "Invalid data format. Missing data start marker '<>' or it's not in the right place"
        )

    if data[2] != TNS_HEADER:
        raise ResultsParsingError("Invalid data format. Missing header row")

    # Convert to polars DataFrame for processing
    df = pl.DataFrame(data[3:], orient="row", schema=TNS_SCHEMA)

    # Process results row by row
    aggregated = []
    formulation_count = 1
    triple_start_indexes = [1, 4, 7]  # Columns for each triplicate set
    
    for row in df.iter_rows():
        # Calculate control value from last three columns
        control_value = statistics.mean(row[10:])
        
        # Process each triplicate set
        for start_index in triple_start_indexes:
            value = statistics.mean(
                row[start_index:start_index + 3]) / control_value
            
            # Validate calculated value
            if value < 10:
                raise ResultsParsingError(
                    "Value {0} for formulation_{1} is less than 10".format(
                        value, formulation_count))
            
            aggregated.append(
                ("formulation_{}".format(formulation_count), value))
            formulation_count += 1

    # Store results in database
    (connection, cursor) = database_helpers.connect_to_database()
    cursor.executemany(
        "INSERT INTO results(experiment_type, formulation_id, calculated_value) VALUES(\"TNS\", ?, ?)",
        aggregated)
    connection.commit()
    cursor.connection.close()

    return aggregated


def parse_zeta_potential_results(filename: str) -> List[Tuple[str, float]]:
    """
    Parses Zeta Potential results from a CSV file.

    The CSV file must contain:
    - Columns: "Measurement Type", "Sample Name", "Zeta Potential (mV)"
    - A control sample named "STD 1"
    - Three measurements per sample
    - All measurements must result in positive values when normalized to control

    Args:
        filename: Path to the CSV file containing Zeta Potential results

    Returns:
        List of tuples containing (formulation_id, calculated_value)

    Raises:
        ResultsParsingError: If measurement counts are incorrect or if normalized values are negative
    """
    # Read and preprocess CSV data
    df = pl.read_csv(filename).select(
        ["Measurement Type", "Sample Name",
         "Zeta Potential (mV)"]).drop_nulls()
    
    # Extract control value
    control_value = df.filter(
        pl.col("Sample Name") == "STD 1").mean()["Zeta Potential (mV)"][0]
    
    # Remove control measurements from dataset
    df = df.filter(pl.col("Sample Name") != "STD 1")

    # Group by sample name for validation and processing
    grouped_by_sample_name = df.group_by("Sample Name")

    # Validate measurement counts
    for (formulation_name, formulation_count
         ) in grouped_by_sample_name.len().sort("Sample Name").iter_rows():
        if formulation_count != 3:
            raise ResultsParsingError("Expected 3 values for {0}, got {1}".format(
                formulation_name, formulation_count))

    # Calculate normalized results
    result = grouped_by_sample_name.agg(
        pl.col("Zeta Potential (mV)").mean() / control_value).rename({
            "Zeta Potential (mV)":
            "calculated_value",
            "Sample Name":
            "formulation_id"
        }).sort("formulation_id")

    # Convert to list of tuples for database storage
    as_dict = result.to_dict(as_series=False)
    as_tuples = list(
        zip(as_dict["formulation_id"], as_dict["calculated_value"]))
    
    # Validate results
    for formulation_id, calculated_value in as_tuples:
        if calculated_value < 0:
            raise ResultsParsingError(
                "Invalid data. Result values should all be positive, but result for {0} is {1}"
                .format(formulation_id, calculated_value))

    # Store results in database
    (connection, cursor) = database_helpers.connect_to_database()
    cursor.executemany(
        "INSERT INTO results(experiment_type, formulation_id, calculated_value) VALUES(\"ZETA_POTENTIAL\", ?, ?)",
        as_tuples)
    connection.commit()
    cursor.connection.close()

    return as_tuples