from openpyxl import load_workbook
import polars as pl
import polars.datatypes
import polars.selectors as cs
import statistics
from datetime import datetime
import sqlite3


def parse_file(filename, upload_timestamp):
    type = None
    try:
        with open(filename, newline='') as f:
            f.read(1024)
        type = "ZETA_POTENTIAL"
    except UnicodeDecodeError:
        type = "TNS"
    if type == "TNS":
        return parse_tns_results(filename, upload_timestamp)
    if type == "ZETA_POTENTIAL":
        return parse_zeta_potential_results(filename, upload_timestamp)
    raise Exception("Could not determine filetype!")


TNS_HEADER = ('<>', 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)

TNS_SCHEMA = {
    "<>": polars.datatypes.String,
    "1": polars.datatypes.UInt64,
    "2": polars.datatypes.UInt64,
    "3": polars.datatypes.UInt64,
    "4": polars.datatypes.UInt64,
    "5": polars.datatypes.UInt64,
    "6": polars.datatypes.UInt64,
    "7": polars.datatypes.UInt64,
    "8": polars.datatypes.UInt64,
    "9": polars.datatypes.UInt64,
    "10": polars.datatypes.UInt64,
    "11": polars.datatypes.UInt64,
    "12": polars.datatypes.UInt64,
}


def connect_to_database(database_path="results.db"):
    connection = sqlite3.connect(database_path)
    cursor = connection.cursor()
    return (connection, cursor)


def parse_tns_results(filename, upload_timestamp):
    wb = load_workbook(filename=filename)
    ws = wb.active

    if ws.dimensions != "A1:M11":
        raise Exception(
            "Invalid data format. The extent of occupied cells should be A1:M11"
        )

    data = list(ws.values)

    if data[0][0] != "Instrument:":
        raise Exception("Invalid data format. 'Instrument:' label is missing")

    instrument = data[0][1]
    if not instrument:
        raise Exception("Invalid data format. Missing instrument name")

    if any(data[0][2:] + data[1]):
        raise Exception(
            "Invalid data format. Cells that should be empty aren't")

    if data[2][0] != "<>":
        raise Exception(
            "Invalid data format. Missing data start marker '<>' or it's not in the right place"
        )

    if data[2] != TNS_HEADER:
        raise Exception("Invalid data format. Missing header row")

    df = pl.DataFrame(data[3:], orient="row", schema=TNS_SCHEMA)
    # Because of the unusual geometry of the results table, the results are parsed by row
    # as polars sees everything in terms of columns
    aggregated = []
    formulation_count = 1
    triple_start_indexes = [1, 4, 7]
    for row in df.iter_rows():
        control_value = statistics.mean(row[10:])
        for start_index in triple_start_indexes:
            value = statistics.mean(
                row[start_index:start_index + 3]) / control_value
            if value < 10:
                raise Exception(
                    "Value {0} for formulation_{1} is less than 10".format(
                        value, formulation_count))
            aggregated.append(
                ("formulation_{}".format(formulation_count), value))
            formulation_count += 1

    (connection, cursor) = connect_to_database()
    cursor.executemany(
        "INSERT INTO results(experiment_type, formulation_id, calculated_value) VALUES(\"TNS\", ?, ?)",
        aggregated)
    connection.commit()
    cursor.connection.close()

    return aggregated


def parse_zeta_potential_results(filename, upload_timestamp):
    df = pl.read_csv(filename).select(
        ["Measurement Type", "Sample Name",
         "Zeta Potential (mV)"]).drop_nulls()
    control_value = df.filter(
        pl.col("Sample Name") == "STD 1").mean()["Zeta Potential (mV)"][0]
    df = df.filter(pl.col("Sample Name") != "STD 1")

    grouped_by_sample_name = df.group_by("Sample Name")

    for (formulation_name, formulation_count
         ) in grouped_by_sample_name.len().sort("Sample Name").iter_rows():
        if formulation_count != 3:
            raise Exception("Expected 3 values for {0}, got {1}".format(
                formulation_name, formulation_count))

    result = grouped_by_sample_name.agg(
        pl.col("Zeta Potential (mV)").mean() / control_value).rename({
            "Zeta Potential (mV)":
            "calculated_value",
            "Sample Name":
            "formulation_id"
        }).sort("formulation_id")

    as_dict = result.to_dict(as_series=False)
    as_tuples = list(
        zip(as_dict["formulation_id"], as_dict["calculated_value"]))
    # Print the result
    for formulation_id, calculated_value in as_tuples:
        if calculated_value < 0:
            raise Exception(
                "Invalid data. Result values should all be positive, but result for {0} is {1}"
                .format(formulation_id, calculated_value))

    (connection, cursor) = connect_to_database()
    cursor.executemany(
        "INSERT INTO results(experiment_type, formulation_id, calculated_value) VALUES(\"ZETA_POTENTIAL\", ?, ?)",
        as_tuples)
    connection.commit()
    cursor.connection.close()

    return as_tuples
